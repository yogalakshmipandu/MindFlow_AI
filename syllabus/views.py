
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from django.db import transaction
import fitz
import re
import json
import os
import requests
from groq import Groq
from dotenv import load_dotenv
from .models import Syllabus, Unit, Topic, TopicLink, TopicDocument
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.drawing.image import Image as ExcelImage

load_dotenv()
client = Groq(api_key=os.getenv("GROQ_API_KEY"))

def extract_text_from_pdf(file):
    """Extract text from PDF file using PyMuPDF."""
    file.seek(0)
    text = ""
    try:
        with fitz.open(stream=file.read(), filetype="pdf") as doc:
            for page in doc:
                text += page.get_text()
    except Exception as e:
        print(f"Error extracting PDF: {e}")
    file.seek(0)
    return text

def extract_units(text):
    """Extract units from syllabus text using regex."""
    unit_pattern = r"(UNIT\s+[IVX\d]+\s*[:\-]?\s*[A-Za-z\s]*)" 
    parts = re.split(unit_pattern, text, flags=re.IGNORECASE)
    units = {}
    for i in range(1, len(parts), 2):
        if i+1 < len(parts):
            unit_name = parts[i].strip()
            unit_content = parts[i+1].strip()
            if unit_name:
                units[unit_name] = unit_content
    return units

# def extract_topics_with_llm(unit_text):
#     """Extract topics from unit text using LLM."""
#     prompt = f"""
# Extract topic names from the syllabus text below.
# Return ONLY a JSON array of topic names, nothing else.

# Example format: ["Topic 1", "Topic 2", "Topic 3"]

# Text:
# {unit_text[:3000]}
# """
#     try:
#         response = client.chat.completions.create(
#             model="llama-3.1-8b-instant",
#             messages=[
#                 {"role": "system", "content": "You extract syllabus topics as a JSON array."},
#                 {"role": "user", "content": prompt}
#             ],
#             temperature=0
#         )
#         content = response.choices[0].message.content.strip()
#         # Try to parse JSON array
#         topics = json.loads(content)
#         return topics if isinstance(topics, list) else []
#     except Exception as e:
#         print(f"Error extracting topics: {e}")
#         return []


def extract_topics_with_llm(unit_text):
    prompt = f"""
Extract only topic names from this syllabus.

Return ONLY a JSON array.

Text:
{unit_text[:3000]}
"""
    try:
        response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[
                {"role": "system", "content": "Return only JSON array of topics."},
                {"role": "user", "content": prompt}
            ],
            temperature=0
        )

        content = response.choices[0].message.content.strip()
        print("RAW RESPONSE:", content)

        match = re.search(r'\[.*\]', content, re.DOTALL)
        if match:
            topics = json.loads(match.group())
            return topics if isinstance(topics, list) else []

        return []

    except Exception as e:
        print(f"Error extracting topics: {e}")
        return []

        
@login_required
@transaction.atomic
def upload_syllabus(request):
    """Handle syllabus PDF upload and extraction."""
    if request.method == "POST":
        pdf_file = request.FILES.get("syllabus")
        title = request.POST.get("title", "Untitled Syllabus")
        
        if not pdf_file:
            messages.error(request, "Please select a PDF file.")
            return redirect("dashboard")
        
        if not pdf_file.name.endswith('.pdf'):
            messages.error(request, "Only PDF files are allowed.")
            return redirect("dashboard")
        
        try:
            # File size check
            if pdf_file.size > 50 * 1024 * 1024:  # 50MB limit
                messages.error(request, "PDF file is too large. Maximum 50MB allowed.")
                return redirect("dashboard")
            
            print(f"[UPLOAD] ===== START UPLOAD =====")
            print(f"[UPLOAD] User: {request.user.username}, File: {pdf_file.name}, Size: {pdf_file.size} bytes")
            
            # STEP 1: Create syllabus record immediately
            print(f"[UPLOAD] Step 1: Creating syllabus record in database...")
            syllabus = Syllabus.objects.create(
                user=request.user,
                title=title,
                pdf_file=pdf_file
            )
            print(f"[UPLOAD] Step 1 SUCCESS: Syllabus ID {syllabus.id} created in DB")
            
            # STEP 2: Extract text from PDF
            print(f"[UPLOAD] Step 2: Extracting text from PDF...")
            text = extract_text_from_pdf(pdf_file)
            print(f"[UPLOAD] Step 2 SUCCESS: Extracted {len(text)} characters")
            
            if not text.strip():
                print(f"[UPLOAD] WARNING: No text extracted from PDF, skipping unit creation")
                messages.success(request, f"Syllabus uploaded successfully (no text content found).")
                return redirect("syllabus_detail", syllabus_id=syllabus.id)
            
            # STEP 3: Extract units
            print(f"[UPLOAD] Step 3: Extracting units...")
            units_dict = extract_units(text)
            print(f"[UPLOAD] Step 3 SUCCESS: Found {len(units_dict)} units")
            
            # STEP 4: Save units
            print(f"[UPLOAD] Step 4: Saving {len(units_dict)} units to database...")
            for idx, (unit_name, unit_content) in enumerate(units_dict.items(), 1):
                unit = Unit.objects.create(
                    syllabus=syllabus,
                    name=unit_name,
                    content=unit_content[:5000]
                )
                print(f"[UPLOAD]   - Unit {idx}/{len(units_dict)}: {unit_name} (ID {unit.id})")
            
            print(f"[UPLOAD] Step 4 SUCCESS: All units saved")
            print(f"[UPLOAD] ===== UPLOAD COMPLETE =====")
            
            messages.success(request, f"Syllabus '{title}' uploaded successfully! Found {len(units_dict)} units.")
            return redirect("syllabus_detail", syllabus_id=syllabus.id)
            
        except Exception as e:
            print(f"[UPLOAD] ===== ERROR OCCURRED =====")
            print(f"[UPLOAD] Exception Type: {type(e).__name__}")
            print(f"[UPLOAD] Exception Message: {str(e)}")
            import traceback
            print(f"[UPLOAD] Traceback:")
            traceback.print_exc()
            print(f"[UPLOAD] ===== END ERROR =====")
            messages.error(request, f"Error processing syllabus: {str(e)}")
            return redirect("dashboard")
    
    return redirect("dashboard")

@login_required
def syllabus_detail(request, syllabus_id):
    """Show syllabus details with units."""
    syllabus = get_object_or_404(Syllabus, id=syllabus_id, user=request.user)
    units = syllabus.units.all()
    return render(request, "syllabus_detail.html", {"syllabus": syllabus, "units": units})

def _safe_excel_text(value):
    return str(value).replace('"', '""') if value is not None else ""

def _build_hyperlink_formula(entries, default_prefix, icon):
    if not entries:
        return None
    formulas = []
    for idx, entry in enumerate(entries, start=1):
        url = entry.get("url") if isinstance(entry, dict) else getattr(entry, "url", "")
        title = entry.get("title", "") if isinstance(entry, dict) else getattr(entry, "title", "")
        if not title:
            title = f"{icon} {default_prefix} {idx}"
        escaped_url = _safe_excel_text(url)
        escaped_title = _safe_excel_text(title)
        formulas.append(f'HYPERLINK("{escaped_url}", "{escaped_title}")')
    if len(formulas) == 1:
        return f"={formulas[0]}"
    return "=" + " & CHAR(10) & ".join(formulas)

def _unit_color(unit_name):
    pastel_colors = [
        "F9E2EE", "E3F6FF", "FFF3BF", "E2F0CB",
        "FDE2F3", "D8E2FF", "FFF1E6", "DCE2D0"
    ]
    index = sum(ord(ch) for ch in str(unit_name)) % len(pastel_colors)
    return pastel_colors[index]

def _normalize_filename(value):
    return re.sub(r"[^A-Za-z0-9_-]+", "_", value).strip("_") or "syllabus_export"

@login_required
def export_syllabus(request, syllabus_id):
    """Export a syllabus as a styled MindFlow-style Excel workbook."""
    syllabus = get_object_or_404(Syllabus, id=syllabus_id, user=request.user)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MindFlow Export"

    headers = ["Syllabus", "Unit", "Topic", "Websites", "YouTube", "Drive", "Text", "Images"]
    sheet.append(headers)

    header_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(wrap_text=True, vertical="top")

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    row_index = 2
    for unit in syllabus.units.all().order_by("id"):
        topics = unit.topics.all().order_by("id")
        for topic in topics:
            links = list(topic.links.all())
            website_links = [link for link in links if link.link_type == "website"]
            youtube_links = [link for link in links if link.link_type == "youtube"]
            text_links = [link for link in links if link.link_type == "doc"]
            image_links = [link for link in links if link.link_type == "image"]
            drive_docs = list(topic.documents.all())

            website_formula = _build_hyperlink_formula(website_links, "Resource", "🔗")
            youtube_formula = _build_hyperlink_formula(youtube_links, "Video", "▶")
            text_formula = _build_hyperlink_formula(text_links, "Doc", "📄")

            drive_entries = []
            for idx, doc in enumerate(drive_docs, start=1):
                doc_url = request.build_absolute_uri(doc.file.url)
                drive_entries.append({"url": doc_url, "title": doc.title or f"📁 File {idx}"})
            drive_formula = _build_hyperlink_formula(drive_entries, "File", "📁")

            image_formula = _build_hyperlink_formula(image_links, "Img", "🖼️")
            row = [
                syllabus.title,
                unit.name,
                topic.name,
                website_formula or "",
                youtube_formula or "",
                drive_formula or "",
                text_formula or "",
                image_formula or ""
            ]
            sheet.append(row)

            fill_color = _unit_color(unit.name)
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            for col in range(1, len(headers) + 1):
                cell = sheet.cell(row=row_index, column=col)
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if col == 3:
                    cell.font = Font(bold=True)

            row_index += 1

    sheet.freeze_panes = sheet["A2"]
    sheet.sheet_view.showGridLines = True

    widths = {
        "A": 22,
        "B": 24,
        "C": 34,
        "D": 32,
        "E": 32,
        "F": 28,
        "G": 28,
        "H": 24,
    }
    for col_letter, width in widths.items():
        sheet.column_dimensions[col_letter].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"{_normalize_filename(syllabus.title)}_mindflow_export.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=\"{filename}\""
    return response

@login_required
def delete_syllabus(request, syllabus_id):
    """Delete a syllabus and all its associated data."""
    syllabus = get_object_or_404(Syllabus, id=syllabus_id, user=request.user)
    
    try:
        syllabus.delete()
        messages.success(request, "Syllabus deleted successfully!")
    except Exception as e:
        messages.error(request, f"Error deleting syllabus: {str(e)}")
    
    return redirect("dashboard")

@login_required
def unit_detail(request, unit_id):
    """Show unit details with topics."""
    unit = get_object_or_404(Unit, id=unit_id, syllabus__user=request.user)
    topics = unit.topics.all()
    return render(request, "unit_detail.html", {"unit": unit, "topics": topics})

@login_required
def extract_topics(request, unit_id):
    """Extract topics for a unit using LLM."""
    unit = get_object_or_404(Unit, id=unit_id, syllabus__user=request.user)
    
    if request.method == "POST":
        topics = extract_topics_with_llm(unit.content)
        
        # Save topics to database
        for topic_name in topics:
            Topic.objects.create(unit=unit, name=topic_name)
        
        messages.success(request, f"Extracted {len(topics)} topics!")
        return redirect("unit_detail", unit_id=unit.id)
    
    return redirect("unit_detail", unit_id=unit.id)

@login_required
def topic_detail(request, topic_id):
    """Show topic details with links and documents."""
    topic = get_object_or_404(Topic, id=topic_id, unit__syllabus__user=request.user)
    links = topic.links.all()
    documents = topic.documents.all()
    return render(request, "topic_detail.html", {"topic": topic, "links": links, "documents": documents})

@login_required
def add_link(request, topic_id):
    """Add a link to a topic."""
    topic = get_object_or_404(Topic, id=topic_id, unit__syllabus__user=request.user)
    
    if request.method == "POST":
        link_type = request.POST.get("link_type")
        url = request.POST.get("url")
        title = request.POST.get("title", "")
        
        if url:
            TopicLink.objects.create(
                topic=topic,
                link_type=link_type,
                url=url,
                title=title
            )
            messages.success(request, "Link added successfully!")
        
        return redirect("topic_detail", topic_id=topic.id)
    
    return redirect("topic_detail", topic_id=topic.id)

@login_required
def delete_link(request, link_id):
    """Delete a link from a topic."""
    link = get_object_or_404(TopicLink, id=link_id, topic__unit__syllabus__user=request.user)
    topic_id = link.topic.id
    link.delete()
    messages.success(request, "Link deleted!")
    return redirect("topic_detail", topic_id=topic_id)

@login_required
def edit_topic_name(request, topic_id):
    """Edit topic name via AJAX."""
    if request.method == "POST":
        topic = get_object_or_404(Topic, id=topic_id, unit__syllabus__user=request.user)
        new_name = request.POST.get("name", "").strip()
        
        if new_name:
            topic.name = new_name
            topic.save()
            return JsonResponse({"success": True, "name": new_name})
        
        return JsonResponse({"success": False, "error": "Name cannot be empty"})
    
    return JsonResponse({"success": False, "error": "Invalid request"})

@login_required
def create_note(request, topic_id):
    """Create and download notes as Word document."""
    if request.method == "POST":
        topic = get_object_or_404(Topic, id=topic_id, unit__syllabus__user=request.user)
        note_content = request.POST.get("note_content", "")
        
        if note_content:
            # Create PDF document
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []
            
            # Add title
            title = Paragraph(f"Notes: {topic.name}", styles['Title'])
            story.append(title)
            story.append(Spacer(1, 12))
            
            # Add content
            content = Paragraph(note_content.replace('\n', '<br/>'), styles['Normal'])
            story.append(content)
            
            doc.build(story)
            buffer.seek(0)
            
            # Return as downloadable file
            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{topic.name}_notes.pdf"'
            return response
        
        messages.error(request, "Please write some content for your notes.")
        return redirect("topic_detail", topic_id=topic.id)
    
    return redirect("topic_detail", topic_id=topic.id)

@login_required
def add_topic(request, unit_id):
    """Add a new topic to a unit."""
    if request.method == "POST":
        unit = get_object_or_404(Unit, id=unit_id, syllabus__user=request.user)
        topic_name = request.POST.get("topic_name", "").strip()
        
        if topic_name:
            Topic.objects.create(unit=unit, name=topic_name)
            messages.success(request, "Topic added successfully!")
        
        return redirect("unit_detail", unit_id=unit.id)
    
    return redirect("unit_detail", unit_id=unit_id)

@login_required
def edit_topic(request, topic_id):
    """Edit a topic's name."""
    if request.method == "POST":
        topic = get_object_or_404(Topic, id=topic_id, unit__syllabus__user=request.user)
        topic_name = request.POST.get("topic_name", "").strip()
        
        if topic_name:
            topic.name = topic_name
            topic.save()
            messages.success(request, "Topic updated successfully!")
        
        return redirect("unit_detail", unit_id=topic.unit.id)
    
    return redirect("unit_detail", unit_id=topic.unit.id)

@login_required
def delete_topic(request, topic_id):
    """Delete a topic."""
    topic = get_object_or_404(Topic, id=topic_id, unit__syllabus__user=request.user)
    unit_id = topic.unit.id
    topic.delete()
    messages.success(request, "Topic deleted!")
    return redirect("unit_detail", unit_id=unit_id)


@login_required
def upload_topic_document(request, topic_id):
    """Upload a document/note for a topic."""
    topic = get_object_or_404(Topic, id=topic_id, unit__syllabus__user=request.user)
    
    if request.method == "POST":
        title = request.POST.get("title", "").strip()
        document_type = request.POST.get("document_type", "notes")
        doc_file = request.FILES.get("document")
        
        if not title:
            title = doc_file.name if doc_file else "Untitled Document"
        
        if doc_file:
            TopicDocument.objects.create(
                topic=topic,
                title=title,
                file=doc_file,
                document_type=document_type
            )
            messages.success(request, "Document uploaded successfully!")
        else:
            messages.error(request, "Please select a file to upload.")
        
        return redirect("topic_detail", topic_id=topic.id)
    
    return redirect("topic_detail", topic_id=topic.id)


@login_required
def delete_topic_document(request, doc_id):
    """Delete a document from a topic."""
    doc = get_object_or_404(TopicDocument, id=doc_id, topic__unit__syllabus__user=request.user)
    topic_id = doc.topic.id
    doc.delete()
    messages.success(request, "Document deleted!")
    return redirect("topic_detail", topic_id=topic_id)

